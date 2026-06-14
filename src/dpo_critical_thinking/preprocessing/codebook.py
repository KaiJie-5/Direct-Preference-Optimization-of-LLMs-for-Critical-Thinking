from __future__ import annotations

import json
import re
import time
from pathlib import Path
from typing import Any


def convert_xlsx_codebook(
    *,
    input_xlsx: Path,
    output_path: Path,
    codebook_id: str,
    codebook_version: str,
    description: str,
    overwrite: bool = False,
) -> dict[str, Any]:
    if output_path.exists() and not overwrite:
        raise FileExistsError(f"Codebook already exists: {output_path}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to convert XLSX codebooks. "
            "Install project dependencies from pyproject.toml."
        ) from exc

    workbook = load_workbook(input_xlsx, data_only=True, read_only=True)
    codes: dict[str, dict[str, Any]] = {}

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_clean_cell(value) for value in rows[0]]
        normalized_headers = [_slugify(header) for header in headers]
        if {"code", "quotes", "example_questions"}.issubset(normalized_headers):
            _read_code_quote_question_sheet(sheet.title, headers, rows[1:], codes)
        elif {"quote", "codes"}.issubset(normalized_headers):
            _read_quote_codes_sheet(sheet.title, headers, rows[1:], codes)

    payload = {
        "codebook_id": codebook_id,
        "codebook_version": codebook_version,
        "source_file": str(input_xlsx),
        "description": description,
        "created_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "codes": list(codes.values()),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return payload


def load_codebook(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if "codes" not in payload or not isinstance(payload["codes"], list):
        raise ValueError(f"Codebook must contain a list field named 'codes': {path}")
    return payload


def _read_code_quote_question_sheet(
    sheet_name: str,
    headers: list[str],
    rows: list[tuple[Any, ...]],
    codes: dict[str, dict[str, Any]],
) -> None:
    code_col = _column_index(headers, "Code")
    quote_col = _column_index(headers, "Quotes")
    question_col = _column_index(headers, "Example Questions")
    current_code_label: str | None = None

    for row in rows:
        code_label = _clean_cell(_get_cell(row, code_col))
        quote = _clean_cell(_get_cell(row, quote_col))
        question = _clean_cell(_get_cell(row, question_col))
        if code_label:
            current_code_label = code_label
        if not current_code_label:
            continue
        entry = _code_entry(codes, current_code_label, sheet_name)
        if quote:
            _append_unique(entry["example_quotes"], quote)
        if question:
            _append_unique(entry["example_reflective_questions"], question)


def _read_quote_codes_sheet(
    sheet_name: str,
    headers: list[str],
    rows: list[tuple[Any, ...]],
    codes: dict[str, dict[str, Any]],
) -> None:
    quote_col = _column_index(headers, "Quote")
    codes_col = _column_index(headers, "Codes")
    current_quote: str | None = None

    for row in rows:
        quote = _clean_cell(_get_cell(row, quote_col))
        code_label = _clean_cell(_get_cell(row, codes_col))
        if quote:
            current_quote = quote
        if not current_quote or not code_label:
            continue
        entry = _code_entry(codes, code_label, sheet_name)
        _append_unique(entry["example_quotes"], current_quote)


def _code_entry(
    codes: dict[str, dict[str, Any]], code_label: str, source_sheet: str
) -> dict[str, Any]:
    code_id = _slugify(code_label)
    if code_id not in codes:
        codes[code_id] = {
            "code_id": code_id,
            "code_label": code_label,
            "definition": None,
            "example_quotes": [],
            "example_reflective_questions": [],
            "source_sheets": [],
        }
    _append_unique(codes[code_id]["source_sheets"], source_sheet)
    return codes[code_id]


def _column_index(headers: list[str], name: str) -> int:
    wanted = _slugify(name)
    for index, header in enumerate(headers):
        if _slugify(header) == wanted:
            return index
    raise ValueError(f"Column {name!r} not found in headers {headers!r}")


def _get_cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _clean_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _append_unique(values: list[str], value: str) -> None:
    if value not in values:
        values.append(value)


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", value.strip().lower()).strip("_")
    return slug or "unnamed_code"
